"""Report content gathering and generation (PDF / Excel / CSV)."""
import csv
import io

from django.core.files.base import ContentFile

from .models import Report

REPORT_META = {
    Report.ReportType.GENERAL: "General Impact Report",
    Report.ReportType.PROGRAMME: "Programme Report",
    Report.ReportType.SURVEY: "Survey Report",
    Report.ReportType.IMPACT: "Impact Report",
    Report.ReportType.KPI: "KPI Report",
    Report.ReportType.PROJECT: "Impact Project Report",
}


def build_context(report: Report) -> dict:
    """Collect all data needed to render a report into a plain dict."""
    from django.utils import timezone

    from apps.impact.analytics import _impact, _monitoring, _overview, _survey

    org = report.organisation
    params = {}
    scope_label = org.name

    if report.programme_id:
        params["programme"] = str(report.programme_id)
        scope_label = report.programme.name
    if report.survey_id:
        params["survey"] = str(report.survey_id)
    if report.impact_project_id:
        params["impact_project"] = str(report.impact_project_id)
        scope_label = report.impact_project.name

    overview = _overview(org, params)
    impact = _impact(org, params)
    monitoring = _monitoring(org, params)
    survey_data = _survey(org, params)

    survey_detail = _survey_detail(report.survey) if report.survey_id else None

    return {
        "title": report.title or REPORT_META.get(report.report_type, "Impact Report"),
        "organisation": org.name,
        "report_type": report.get_report_type_display(),
        "file_format": report.get_file_format_display(),
        "scope": scope_label,
        "generated_at": timezone.now().strftime("%d %B %Y"),
        "overview": overview,
        "kpis": impact["kpis"],
        "baseline_endline": impact["baseline_endline"],
        "challenges": monitoring,
        "surveys": survey_data,
        "survey_detail": survey_detail,
        "programme": report.programme,
        "impact_project": report.impact_project,
    }


def _survey_detail(survey):
    """Per-question response breakdown for survey reports."""
    rows = []
    for question in survey.questions.all().order_by("order"):
        breakdown = {}
        for answer in question.answers.all():
            value = answer.value
            key = value if isinstance(value, str) else (
                ", ".join(str(v) for v in value) if isinstance(value, list) else str(value)
            )
            breakdown[key] = breakdown.get(key, 0) + 1
        rows.append(
            {
                "question": question.question,
                "type": question.get_question_type_display(),
                "required": question.required,
                "breakdown": dict(sorted(breakdown.items(), key=lambda item: -item[1])),
                "total_answers": question.answers.count(),
            }
        )
    return {
        "title": survey.title,
        "responses": survey.responses.count(),
        "rows": rows,
    }


def generate_report(report: Report):
    """Generate the report file and mark it READY or FAILED."""
    context = build_context(report)
    try:
        if report.file_format == Report.FileFormat.PDF:
            content = _render_pdf(context)
        elif report.file_format == Report.FileFormat.EXCEL:
            content = _render_excel(context)
        else:
            content = _render_csv(context)
        report.file.save(report.filename, ContentFile(content), save=False)
        report.error_message = ""
        report.status = Report.Status.READY
        report.save(update_fields=["file", "status", "error_message"])
    except Exception as exc:  # pragma: no cover - defensive guard
        report.status = Report.Status.FAILED
        report.error_message = str(exc)[:1000]
        report.save(update_fields=["status", "error_message"])
def _render_pdf(context):
    """Professionally formatted PDF report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.9 * inch,
        leftMargin=0.9 * inch,
        topMargin=0.9 * inch,
        bottomMargin=0.9 * inch,
        title=context["title"],
    )

    title_style = ParagraphStyle(
        "Title", fontName="Helvetica-Bold", fontSize=22, leading=26, spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        fontName="Helvetica",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#666666"),
    )
    heading_style = ParagraphStyle(
        "Heading",
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        spaceBefore=18,
        spaceAfter=6,
    )
    body_style = ParagraphStyle("Body", fontName="Helvetica", fontSize=10, leading=14)

    overview = context["overview"]
    elements = [
        Paragraph(context["organisation"].upper(), subtitle_style),
        Paragraph(context["title"], title_style),
        Paragraph(
            f"{context['report_type']} • Scope: {context['scope']} • {context['generated_at']}",
            subtitle_style,
        ),
        Spacer(1, 0.2 * inch),
    ]

    # 1. Overview
    elements.append(Paragraph("Programme Overview", heading_style))
    overview_rows = [
        ["Metric", "Value"],
        ["Participants reached", overview.get("participants_reached", 0)],
        ["Enrolment", overview.get("enrolment", 0)],
        ["Survey responses", overview.get("survey_responses", 0)],
        ["Survey response rate", f"{overview.get('survey_response_rate', 0)}%"],
        ["Completion rate", f"{overview.get('completion_rate', 0)}%"],
        ["Active programmes", overview.get("active_programmes", 0)],
        ["Active surveys", overview.get("active_surveys", 0)],
    ]
    elements.append(_styled_table(overview_rows))

    # 2. KPIs
    elements.append(Paragraph("KPIs and Progress", heading_style))
    kpi_rows = [["KPI", "Baseline", "Current", "Target", "Progress", "Status"]]
    for kpi in context["kpis"]:
        kpi_rows.append(
            [
                kpi["kpi"],
                kpi.get("baseline") if kpi.get("baseline") is not None else "-",
                kpi.get("current"),
                kpi.get("target") if kpi.get("target") is not None else "-",
                f"{kpi.get('progress_percentage', 0)}%",
                kpi.get("status", ""),
            ]
        )
    elements.append(_styled_table(kpi_rows))

    # 3. Baseline vs Endline
    if context["baseline_endline"]:
        elements.append(Paragraph("Baseline vs Endline", heading_style))
        be_rows = [["KPI", "Baseline", "Endline", "Change"]]
        for item in context["baseline_endline"]:
            be_rows.append(
                [
                    item["kpi"],
                    item.get("baseline") if item.get("baseline") is not None else "-",
                    item.get("endline") if item.get("endline") is not None else "-",
                    item.get("change") if item.get("change") is not None else "-",
                ]
            )
        elements.append(_styled_table(be_rows))

    # 4. Challenges
    elements.append(Paragraph("Monitoring & Challenges", heading_style))
    challenges = context["challenges"]
    challenge_rows = [
        ["Status", "Count"],
        ["Open", challenges.get("open", 0)],
        ["In progress", challenges.get("in_progress", 0)],
        ["Resolved", challenges.get("resolved", 0)],
        ["Total", challenges.get("total_challenges", 0)],
        ["Resolution rate", f"{challenges.get('resolution_rate', 0)}%"],
    ]
    elements.append(_styled_table(challenge_rows))

    # 5. Survey detail (when applicable)
    if context["survey_detail"]:
        elements.append(Paragraph(f"Survey: {context['survey_detail']['title']}", heading_style))
        elements.append(Paragraph(f"Total responses: {context['survey_detail']['responses']}", body_style))
        for row in context["survey_detail"]["rows"]:
            elements.append(Paragraph(f"{row['question']} ({row['type']})", body_style))
            if row["breakdown"]:
                breakdown_rows = [["Answer", "Count"]]
                for answer, count_value in row["breakdown"].items():
                    breakdown_rows.append([str(answer)[:120], count_value])
                elements.append(_styled_table(breakdown_rows))

    elements.append(Spacer(1, 0.3 * inch))
    elements.append(
        Paragraph(
            "Generated by the Voice of a Girl Impact Platform. This report contains "
            "aggregated programme data for programme management and donor reporting.",
            body_style,
        )
    )

    doc.build(elements)
    return buffer.getvalue()


def _styled_table(rows):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#312e81")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8.5),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#eef2ff")]),
                ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#c7d2fe")),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]
        )
    )
    return table
def _render_excel(context):
    """Excel workbook with meaningful worksheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")

    def style_header(ws, cols):
        for index, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=index, value=col)
            cell.font = header_font
            cell.fill = header_fill

    # Sheet 1: Overview
    ws = wb.active
    ws.title = "Overview"
    overview = context["overview"]
    overview_data = [
        ("Metric", "Value"),
        ("Organisation", context["organisation"]),
        ("Scope", context["scope"]),
        ("Report generated", context["generated_at"]),
        ("Participants reached", overview.get("participants_reached", 0)),
        ("Enrolment", overview.get("enrolment", 0)),
        ("Survey responses", overview.get("survey_responses", 0)),
        ("Survey response rate (%)", overview.get("survey_response_rate", 0)),
        ("Completion rate (%)", overview.get("completion_rate", 0)),
        ("Active programmes", overview.get("active_programmes", 0)),
        ("Active surveys", overview.get("active_surveys", 0)),
    ]
    for row_index, row in enumerate(overview_data, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
    style_header(ws, ["Metric", "Value"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Sheet 2: KPIs
    ws2 = wb.create_sheet("KPIs")
    style_header(ws2, ["KPI", "Baseline", "Current", "Target", "Endline", "Progress (%)", "Status"])
    for row_index, kpi in enumerate(context["kpis"], start=2):
        values = [
            kpi["kpi"],
            kpi.get("baseline") or 0,
            kpi.get("current") or 0,
            kpi.get("target") or 0,
            kpi.get("endline") or 0,
            kpi.get("progress_percentage", 0),
            kpi.get("status", ""),
        ]
        for col_index, value in enumerate(values, start=1):
            ws2.cell(row=row_index, column=col_index, value=value)
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 3: Survey Responses
    ws3 = wb.create_sheet("Survey Responses")
    style_header(ws3, ["Survey", "Responses"])
    for row_index, item in enumerate(context["surveys"].get("responses_per_survey", []), start=2):
        ws3.cell(row=row_index, column=1, value=item["survey__title"])
        ws3.cell(row=row_index, column=2, value=item["count"])
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 12

    # Sheet 4: Challenges
    ws4 = wb.create_sheet("Challenges")
    style_header(ws4, ["Category", "Count"])
    for row_index, item in enumerate(context["challenges"].get("challenges_by_category", []), start=2):
        ws4.cell(row=row_index, column=1, value=item["category"])
        ws4.cell(row=row_index, column=2, value=item["count"])
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _render_csv(context):
    """Comma-separated export suitable for further analysis."""
    buffer_text = io.StringIO()
    writer = csv.writer(buffer_text)
    writer.writerow(["Voice of a Girl Impact Platform - Report"])
    writer.writerow(["Organisation", context["organisation"]])
    writer.writerow(["Scope", context["scope"]])
    writer.writerow(["Report type", context["report_type"]])
    writer.writerow(["Generated", context["generated_at"]])
    writer.writerow([])

    overview = context["overview"]
    writer.writerow(["Metric", "Value"])
    for key, value in overview.items():
        writer.writerow([key.replace("_", " ").title(), value])
    writer.writerow([])

    writer.writerow(["KPI", "Baseline", "Current", "Target", "Endline", "Progress (%)", "Status"])
    for kpi in context["kpis"]:
        writer.writerow(
            [
                kpi["kpi"],
                kpi.get("baseline") or 0,
                kpi.get("current") or 0,
                kpi.get("target") or 0,
                kpi.get("endline") or 0,
                kpi.get("progress_percentage", 0),
                kpi.get("status", ""),
            ]
        )
    writer.writerow([])
    writer.writerow(["Challenge Category", "Count"])
    for item in context["challenges"].get("challenges_by_category", []):
        writer.writerow([item["category"], item["count"]])

    output = buffer_text.getvalue()
    return output.encode("utf-8-sig")